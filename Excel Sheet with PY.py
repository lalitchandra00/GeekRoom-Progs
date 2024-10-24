# Opening and going through cells in excel sheet and updating it
import openpyxl as xl
excel = xl.load_workbook("XlSupport.xlsx")
sheet = excel["Sheet1"]

'''Two ways to go to cells in excel'''
cell1 = sheet["A1"]
cell2 = sheet.cell(1,1)
print(cell2.value)
print(cell1.value)

'''Total no of rows and columns'''
print(sheet.max_column)
print(sheet.max_row)

print("")
print("")
print("")

'''To print all the data in excel sheet'''
for row in range(1,sheet.max_row + 1):
    for column in range (1,sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value,end = " ")
    print("\n")


'''To update data in sheet'''
for row in range(2,sheet.max_row+1):
            cell = sheet.cell(row,4)
            corrected = cell.value *10
            corrected_cell = sheet.cell(row,5)
            corrected_cell.value =corrected

            new_rowname = "corrected"
            corrected_cell = sheet.cell(1,5)
            corrected_cell.value = new_rowname

excel.save("XlSupport1.xlsx")


from openpyxl.chart import BarChart, Reference
values = Reference(sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col = 4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,"F2")
excel.save("XlSupport1.xlsx")